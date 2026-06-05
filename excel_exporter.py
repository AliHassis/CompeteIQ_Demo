"""
Excel exporter — creates a multi-sheet professional report.
Sheets: Overview, Price Analysis, Reviews, SEO Keywords, Opportunities
"""
from __future__ import annotations
import io
from datetime import datetime
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter


# ── Colors ──────────────────────────────────────────────────────────────────
DARK_BLUE = "1a1f3a"
ACCENT    = "4f8ef7"
GREEN     = "22c55e"
RED       = "ef4444"
YELLOW    = "f59e0b"
LIGHT_BG  = "f0f4ff"
WHITE     = "FFFFFF"


def _header_fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)

def _thin_border() -> Border:
    s = Side(style="thin", color="d1d5db")
    return Border(left=s, right=s, top=s, bottom=s)

def _style_header_row(ws, row: int, n_cols: int, fill_color: str = DARK_BLUE):
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = _header_fill(fill_color)
        cell.font = Font(bold=True, color=WHITE, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _thin_border()

def _style_data_row(ws, row: int, n_cols: int, even: bool = False):
    fill = PatternFill("solid", fgColor=LIGHT_BG if even else WHITE)
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _thin_border()
        cell.font = Font(size=10)

def _autofit(ws, min_w: int = 12, max_w: int = 45):
    for col_cells in ws.columns:
        length = max(
            len(str(c.value or "")) for c in col_cells
        )
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = \
            min(max(length + 2, min_w), max_w)


# ── Sheet builders ──────────────────────────────────────────────────────────

def _build_overview(ws, products: list[dict]):
    ws.sheet_view.rightToLeft = True
    headers = ["المنصة", "عنوان المنتج", "السعر", "العملة", "التقييم",
               "عدد المراجعات", "التوفر", "البائع", "الرابط"]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))
    ws.row_dimensions[1].height = 30

    for i, p in enumerate(products, start=2):
        ws.append([
            p.get("platform", "N/A"),
            p.get("title", "N/A")[:100],
            p.get("price"),
            p.get("currency", "N/A"),
            p.get("rating"),
            p.get("reviews_count"),
            p.get("availability", "N/A"),
            p.get("seller", "N/A"),
            p.get("url", ""),
        ])
        _style_data_row(ws, i, len(headers), even=(i % 2 == 0))
        # Hyperlink for URL
        cell = ws.cell(row=i, column=9)
        if cell.value and cell.value.startswith("http"):
            cell.hyperlink = cell.value
            cell.value = "🔗 رابط المنتج"
            cell.font = Font(color=ACCENT, underline="single", size=10)

    _autofit(ws)


def _build_price_sheet(ws, price_analysis: dict):
    ws.sheet_view.rightToLeft = True
    # Summary row
    ws["A1"] = "📊 ملخص الأسعار"
    ws["A1"].font = Font(bold=True, size=14, color=DARK_BLUE)
    ws.merge_cells("A1:F1")

    summary_headers = ["المتوسط", "الأدنى", "الأعلى", "الفجوة السعرية", "العملة"]
    summary_vals = [
        price_analysis.get("average"),
        price_analysis.get("min"),
        price_analysis.get("max"),
        price_analysis.get("spread"),
        price_analysis.get("currency", "N/A"),
    ]
    for col, (h, v) in enumerate(zip(summary_headers, summary_vals), start=1):
        ws.cell(row=2, column=col, value=h).font = Font(bold=True, color=DARK_BLUE)
        ws.cell(row=3, column=col, value=v)
    _style_header_row(ws, 2, len(summary_headers), fill_color=ACCENT)
    _style_data_row(ws, 3, len(summary_headers))

    # Detailed table
    ws.append([])  # blank
    headers = ["المنصة", "المنتج", "السعر", "الفرق عن المتوسط", "الفرق %", "الموقع التنافسي"]
    ws.append(headers)
    _style_header_row(ws, 5, len(headers))

    for i, p in enumerate(price_analysis.get("products", []), start=6):
        row = [
            p.get("platform", "N/A"),
            p.get("title", "N/A")[:80],
            p.get("price"),
            p.get("vs_avg"),
            f"{p.get('vs_avg_pct', 0)}%",
            p.get("position", "N/A"),
        ]
        ws.append(row)
        _style_data_row(ws, i, len(headers), even=(i % 2 == 0))
        # Color position cell
        pos_cell = ws.cell(row=i, column=6)
        pos = p.get("position", "")
        if "الأرخص" in pos:
            pos_cell.fill = _header_fill(GREEN)
            pos_cell.font = Font(color=WHITE, bold=True, size=10)
        elif "الأغلى" in pos:
            pos_cell.fill = _header_fill(RED)
            pos_cell.font = Font(color=WHITE, bold=True, size=10)

    _autofit(ws)


def _build_reviews_sheet(ws, review_analysis: dict):
    ws.sheet_view.rightToLeft = True
    ws["A1"] = "⭐ تحليل التقييمات والمراجعات"
    ws["A1"].font = Font(bold=True, size=14, color=DARK_BLUE)

    headers = ["الترتيب", "المنصة", "المنتج", "التقييم", "عدد المراجعات", "درجة الثقة الاجتماعية"]
    ws.append([""])
    ws.append(headers)
    _style_header_row(ws, 3, len(headers))

    for i, p in enumerate(review_analysis.get("ranked", []), start=4):
        ws.append([
            i - 3,
            p.get("platform", "N/A"),
            p.get("title", "N/A")[:80],
            p.get("rating"),
            p.get("reviews_count"),
            p.get("social_proof_score"),
        ])
        _style_data_row(ws, i, len(headers), even=(i % 2 == 0))

    _autofit(ws)


def _build_seo_sheet(ws, seo_analysis: dict):
    ws.sheet_view.rightToLeft = True
    ws["A1"] = "🔍 تحليل الكلمات المفتاحية (SEO)"
    ws["A1"].font = Font(bold=True, size=14, color=DARK_BLUE)

    headers = ["الكلمة المفتاحية المشتركة", "عدد المنتجات التي تستخدمها"]
    ws.append([""])
    ws.append(headers)
    _style_header_row(ws, 3, len(headers))

    for i, kw in enumerate(seo_analysis.get("top_shared_keywords", []), start=4):
        ws.append([kw["keyword"], kw["count"]])
        _style_data_row(ws, i, 2, even=(i % 2 == 0))

    # Title length section
    ws.append([])
    ws.append(["عنوان المنتج", "طول العنوان (حرف)"])
    row_n = ws.max_row
    _style_header_row(ws, row_n, 2, fill_color=ACCENT)
    for p in seo_analysis.get("title_lengths", []):
        row_n += 1
        ws.append([p["title"], p["length"]])
        _style_data_row(ws, row_n, 2, even=(row_n % 2 == 0))

    _autofit(ws)


def _build_opportunities_sheet(ws, opportunities: list[str]):
    ws.sheet_view.rightToLeft = True
    ws["A1"] = "💡 الفرص التنافسية والتوصيات"
    ws["A1"].font = Font(bold=True, size=16, color=DARK_BLUE)
    ws.merge_cells("A1:B1")

    ws.append(["#", "الفرصة / التوصية"])
    _style_header_row(ws, 2, 2)

    for i, opp in enumerate(opportunities, start=3):
        ws.append([i - 2, opp])
        _style_data_row(ws, i, 2, even=(i % 2 == 0))
        ws.cell(row=i, column=2).font = Font(size=11)

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 70


# ── Public API ───────────────────────────────────────────────────────────────

def export_to_excel(
    products: list[dict],
    price_analysis: dict,
    review_analysis: dict,
    seo_analysis: dict,
    opportunities: list[str],
) -> bytes:
    """
    Build a styled multi-sheet Excel report and return as bytes.
    """
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # Write placeholder DataFrames to create sheets
        for name in ["نظرة_عامة", "تحليل_الأسعار", "التقييمات", "SEO_الكلمات", "الفرص"]:
            pd.DataFrame().to_excel(writer, sheet_name=name, index=False)

        wb = writer.book

        _build_overview(wb["نظرة_عامة"], products)
        _build_price_sheet(wb["تحليل_الأسعار"], price_analysis)
        _build_reviews_sheet(wb["التقييمات"], review_analysis)
        _build_seo_sheet(wb["SEO_الكلمات"], seo_analysis)
        _build_opportunities_sheet(wb["الفرص"], opportunities)

        # Add metadata sheet
        meta = wb.create_sheet("معلومات_التقرير")
        meta["A1"] = "🕵️ CompeteIQ — تقرير تحليل المنافسين"
        meta["A1"].font = Font(bold=True, size=16, color=DARK_BLUE)
        meta["A3"] = f"تاريخ التقرير: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        meta["A4"] = f"عدد المنتجات المحللة: {len(products)}"
        meta["A5"] = "تم الإنشاء بواسطة: CompeteIQ"

    output.seek(0)
    return output.read()
